from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden, HttpResponse, JsonResponse
from .forms import ProdutorForm, TerrenoForm, TalhaoForm, UserForm
from django.contrib.auth.decorators import login_required
from .models import Coordenador, Produtor, Terreno, Talhao
from sistema.utils import is_coordenador_or_superuser
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .serializer import ProdutorSerializer
from sistema.utils import importar_servicos_da_api
from datetime import datetime
from django.contrib.auth.models import User
from django.contrib.auth import authenticate


# @api_view(['POST'])
# def api_autenticar_produtor(request):
#     email = request.data.get('email')
#     password = request.data.get('password')

#     if not email or not password:
#         return JsonResponse({'erro': 'E-mail e senha são obrigatórios.'}, status=status.HTTP_400_BAD_REQUEST)

#     # Use o e-mail como `username` na função authenticate.
#     authenticated_user = authenticate(username=email, password=password)

#     if authenticated_user is not None:
#         # A autenticação já está validada, você não precisa fazer mais nada aqui.
#         return JsonResponse({'autenticado': True, 'email': authenticated_user.email, 'nome': authenticated_user.first_name}, status=status.HTTP_200_OK)
#     else:
#         return JsonResponse({'autenticado': False, 'mensagem': 'Credenciais inválidas.'}, status=status.HTTP_401_UNAUTHORIZED)
    
@api_view(['POST'])
def api_autenticar_produtor(request):
    email = request.data.get('email')
    password = request.data.get('password')

    if not email or not password:
        return JsonResponse({'autenticado': False}, status=status.HTTP_400_BAD_REQUEST)

    user = authenticate(username=email, password=password)

    if user is not None:
        return JsonResponse({'autenticado': True}, status=status.HTTP_200_OK)
    return JsonResponse({'autenticado': False}, status=status.HTTP_401_UNAUTHORIZED)

def index(request):
    return render(request,'sistema/welcome.html')

@login_required
def listaProd(request):
    produtores = Produtor.objects.all()
    coordenador = Coordenador.objects.get(user=request.user)

    contexto = {
        "produtores": produtores,
        'coordenador': coordenador
    }

    return render(request,"coordenador/listaProdutores.html",contexto)


@login_required
def criarProdutor(request): 
    if request.method == "POST":
        form = ProdutorForm(request.POST)
        user_form = UserForm(request.POST)
        if form.is_valid() and user_form.is_valid():            
            user = user_form.save(commit=False)
            user.set_password(user.password)
            user.save()
            produtor = form.save(commit=False)
            produtor.coordenador = Coordenador.objects.get(user=request.user)
            produtor.cidade = request.POST.get('cidade')
            produtor.user = user
            produtor.save()
            messages.success(request, "Produtor cadastrado com sucesso!")
            return redirect('coordenador:Lista_Produtores')
            # redirecionar ou mostrar mensagem de sucesso
    else:
        form = ProdutorForm()
        user_form = UserForm()
        
    return render(request, 'coordenador/adicionarProdutor.html', {'form': form, 'user_form': user_form})

@login_required
def produtor_atualiza(request, id):
    produtor = get_object_or_404(Produtor, id=id)
    user = produtor.user

    if not is_coordenador_or_superuser(request.user, produtor):
        return HttpResponseForbidden("Apenas coordenadores podem acessar esse link.")

    if request.method == "POST":
        form = ProdutorForm(request.POST, instance=produtor)
        user_form = UserForm(request.POST, instance=produtor)
        if form.is_valid() and user_form.is_valid():
            produtor_atualizado = form.save(commit=False)
            usuario_atualizado = user_form.save(commit=False)
            produtor.cidade = request.POST.get('cidade')
            nova_senha = request.POST.get('password')

            if nova_senha:
                usuario_atualizado.set_password(nova_senha)

            usuario_atualizado.save()
            usuario_atualizado.save()
            produtor_atualizado.save()
            messages.success(request, "Produtor salvo!")
            return redirect('coordenador:Lista_Produtores')
        else:
            messages.error(request, "Erro ao salvar. Verifique os campos!")
    else:
        form = ProdutorForm(instance=produtor)
        user_form = UserForm(instance=user)
    return render(request, 'coordenador/produtor-atualiza.html', {
        'form': form,
        'user_form': user_form,
        'produtor': produtor,
    })


@login_required
def produtor_exclui(request, id):
    produtor = get_object_or_404(Produtor, id=id)

    if not is_coordenador_or_superuser(request.user, produtor):
        return HttpResponseForbidden("Você não tem autorização para acessar essa página.")

    if request.method == "GET":
        produtor.delete()
        messages.success(request, "Produtor excluído!")
        return redirect('coordenador:Lista_Produtores')
    
@login_required
def produtor_detalhe(request, id):
    produtor = get_object_or_404(Produtor, id=id)

    if not is_coordenador_or_superuser(request.user, produtor):
        return HttpResponseForbidden("Você não tem autorização para acessar essa página.")
    
    terrenos = Terreno.objects.filter(produtor=produtor).order_by('nome')

    return render(request, 'coordenador/produtor-detalhe.html', {
        'produtor': produtor,
        'terrenos': terrenos,
    })

@login_required
def terreno_cria(request, id):
    produtor = get_object_or_404(Produtor, id=id)

    if not is_coordenador_or_superuser(request.user, produtor):
        return HttpResponseForbidden("Você não tem autorização para acessar essa página.")

    if request.method == "POST":
        print("Entrou no POST")
        form = TerrenoForm(request.POST)

        if form.is_valid():
            terreno = form.save(commit=False)
            terreno.produtor = produtor
            terreno.cidade = request.POST.get('cidade')
            terreno.save()
            messages.success(request, "Terreno criado com sucesso!")
            return redirect('coordenador:produtor_detalhe', id=produtor.id)
    else:
        form = TerrenoForm()

    return render(request, 'coordenador/terreno-cria.html', {'form': form, 'produtor': produtor})

@login_required
def terreno_detalhe(request, id):
    terreno = get_object_or_404(Terreno, id=id)
    
    if not is_coordenador_or_superuser(request.user, terreno.produtor):
        return HttpResponseForbidden("Você não tem autorização para acessar essa página.")

    talhoes = Talhao.objects.filter(terreno=terreno).order_by('nome')

    return render(request, 'coordenador/terreno-detalhe.html', {'terreno':terreno, 'talhoes':talhoes})

@login_required
def terreno_atualiza(request, id):
    terreno = get_object_or_404(Terreno, id=id)

    if not is_coordenador_or_superuser(request.user, terreno.produtor):
        return HttpResponseForbidden("Você não tem autorização para acessar essa página.")
    
    if request.method == "POST":
        form = TerrenoForm(request.POST, instance=terreno)
        if form.is_valid():
            terreno_atualizado = form.save(commit=False)
            terreno.cidade = request.POST.get('cidade')
            terreno_atualizado.save()
            messages.success(request, "Terreno salvo!")
            return redirect('coordenador:terreno_detalhe', id=terreno.id)
        else:
            messages.error(request, "Erro ao salvar. Verifique os campos!")
    else:
        form = TerrenoForm(instance=terreno)

    return render(request, 'coordenador/terreno-atualiza.html', {
        'form': form,
        'terreno': terreno,
    })

@login_required
def terreno_exclui(request, id):
    terreno = get_object_or_404(Terreno, id=id)

    if not is_coordenador_or_superuser(request.user, terreno.produtor):
        return HttpResponseForbidden("Você não tem autorização para acessar essa página.")

    if request.method == "GET":
        terreno.delete()
        messages.success(request, "Terreno excluído!")
        return redirect('coordenador:produtor_detalhe', id=terreno.produtor.id)
    
@login_required
def talhao_cria(request, id):
    terreno = get_object_or_404(Terreno, id=id)

    if not is_coordenador_or_superuser(request.user, terreno.produtor):
        return HttpResponseForbidden("Você não tem autorização para acessar essa página.")

    if request.method == "POST":
        print("Entrou no POST")
        form = TalhaoForm(request.POST)

        if form.is_valid():
            talhao = form.save(commit=False)
            talhao.terreno = terreno
            talhao.cidade = request.POST.get('cidade')
            talhao.save()
            messages.success(request, "Terreno criado com sucesso!")
            return redirect('coordenador:terreno_detalhe', id=terreno.id)
    else:
        form = TalhaoForm()

    return render(request, 'coordenador/talhao-cria.html', {'form': form, 'terreno': terreno})

@login_required
def talhao_detalhe(request, id):
    talhao = get_object_or_404(Talhao, id=id)
    
    if not is_coordenador_or_superuser(request.user, talhao.terreno.produtor):
        return HttpResponseForbidden("Você não tem autorização para acessar essa página.")

    return render(request, 'coordenador/talhao-detalhe.html', {'talhao':talhao})

@login_required
def talhao_atualiza(request, id):
    talhao = get_object_or_404(Talhao, id=id)

    if not is_coordenador_or_superuser(request.user, talhao.terreno.produtor):
        return HttpResponseForbidden("Você não tem autorização para acessar essa página.")
    
    if request.method == "POST":
        form = TalhaoForm(request.POST, instance=talhao)
        if form.is_valid():
            talhao_atualizado = form.save(commit=False)
            talhao.cidade = request.POST.get('cidade')
            talhao_atualizado.save()
            messages.success(request, "Talhão salvo!")
            return redirect('coordenador:talhao_detalhe', id=talhao.id)
        else:
            messages.error(request, "Erro ao salvar. Verifique os campos!")
    else:
        form = TalhaoForm(instance=talhao)

    return render(request, 'coordenador/talhao-atualiza.html', {
        'form': form,
        'talhao': talhao,
    })

@login_required
def talhao_exclui(request, id):
    talhao = get_object_or_404(Talhao, id=id)

    if not is_coordenador_or_superuser(request.user, talhao.terreno.produtor):
        return HttpResponseForbidden("Você não tem autorização para acessar essa página.")

    if request.method == "GET":
        talhao.delete()
        messages.success(request, "Terreno excluído!")
        return redirect('coordenador:terreno_detalhe', id=talhao.terreno.produtor.id)
    
def excelGera(request):
    wb = Workbook()
    
    # Aba principal: Produtores
    ws_produtores = wb.active
    ws_produtores.title = "Produtores"
    ws_produtores.append([
        "Nome", "Telefone", "CPF", "Cidade", 
        "CAF", "Validade CAF", "Coordenador"
    ])

    for produtor in Produtor.objects.select_related("coordenador").all():
        ws_produtores.append([
            produtor.nome,
            produtor.telefone,
            produtor.cpf,
            produtor.cidade,
            produtor.caf,
            produtor.validade_caf.strftime('%d/%m/%Y'),
            produtor.coordenador.user.first_name
        ])

    # Abas separadas para cada terreno
    terrenos = Terreno.objects.select_related("produtor").prefetch_related("talhao_set").all()

    for terreno in terrenos:
        # Nome da aba: limitar a 31 caracteres (limite do Excel)
        aba_nome = f"{terreno.produtor.nome[:10]}_{terreno.nome[:20]}"
        aba_nome = aba_nome[:31]

        ws = wb.create_sheet(title=aba_nome)

        # Cabeçalhos
        ws.append([
            f"Terreno: {terreno.nome}",
        ])
        ws.append([
            "Cidade", "Produtor", "Talhão", "Cidade Talhão", 
            "Data Certificação", "Nº Plantas", "Data Plantio", 
            "Variedade", "Área", "Validade CAF"
        ])

        talhoes = Talhao.objects.filter(terreno=terreno)
        for talhao in talhoes:
            ws.append([
                terreno.cidade,
                terreno.produtor.nome,
                talhao.nome,
                talhao.terreno.cidade,
                talhao.data_certificacao.strftime('%d/%m/%Y'),
                talhao.numero_plantas,
                talhao.data_plantio.strftime('%d/%m/%Y'),
                talhao.variedade,
                talhao.area,
                talhao.validade_caf.strftime('%d/%m/%Y')
            ])

    # Gerar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=relatorio_produtores.xlsx'
    wb.save(response)
    return response

def exportar_servicos_excel(request):
    resultado = importar_servicos_da_api()
    if not resultado["success"]:
        return HttpResponse("Erro ao buscar dados da API", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Serviços API"

    headers = ["Produtor", "Talhão", "Data", "Valor Serviço", "Tipos de Serviço", "Produtos", "Trabalhadores"]
    ws.append(headers)

    for servico in resultado["dados"]:
        produtor = servico["proprietario"]["username"]
        talhao = servico["talhao"]
        data = servico["data"]
        valor_servico = servico["valor_servico"]
        
        # Transforma listas compostas em strings legíveis
        tipos = ", ".join(servico.get("servico_tipo", []))
        produtos = ", ".join([
            f'{p["nome"]} ({p["qtde"]} {p["unidade"]}) - R${p["valor"]}'
            for p in servico.get("produtos", [])
        ])
        trabalhadores = ", ".join([
            t["nome"] for t in servico.get("trabalhadores", [])
        ])

        ws.append([produtor, talhao, data, valor_servico, tipos, produtos, trabalhadores])

    # Monta a resposta HTTP com o Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="servicos_api.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_excel_coordenador(request):
    # Garante que o usuário é coordenador ou superuser
    if not hasattr(request.user, 'coordenador') and not request.user.is_superuser:
        return HttpResponse("Acesso não autorizado", status=403)
    
    coordenador = request.user.coordenador

    wb = Workbook()

    # ===== ABA PRINCIPAL: PRODUTORES =====
    ws_produtores = wb.active
    ws_produtores.title = "Produtores"
    ws_produtores.append([
        "Nome", "Telefone", "CPF", "Cidade",
        "CAF", "Validade CAF", "Coordenador"
    ])

    bold = Font(bold=True)
    for cell in ws_produtores["1:1"]:
        cell.font = bold

    produtores = Produtor.objects.filter(coordenador=coordenador).select_related("coordenador", "user")
    for produtor in produtores:
        ws_produtores.append([
            produtor.user.first_name,
            produtor.telefone,
            produtor.cpf,
            produtor.cidade,
            produtor.caf,
            produtor.validade_caf.strftime('%d/%m/%Y'),
            produtor.coordenador.user.first_name
        ])

    # ===== ABAS PARA TERRENOS DESTE COORDENADOR =====
    terrenos = Terreno.objects.filter(produtor__coordenador=coordenador)\
        .select_related('produtor')\
        .prefetch_related('talhoes')

    for terreno in terrenos:
        nome_produtor = terreno.produtor.user.first_name[:10]
        nome_terreno = terreno.nome[:20]
        aba_nome = f"{nome_produtor}_{nome_terreno}"[:31]

        ws = wb.create_sheet(title=aba_nome)
        ws.append([f"Terreno: {terreno.nome}"])
        ws.append([
            "Cidade", "Produtor", "Talhão", "Cidade Talhão",
            "Data Certificação", "Nº Plantas", "Data Plantio",
            "Variedade", "Área", "Validade CAF"
        ])

        for cell in ws["2:2"]:
            cell.font = bold
            cell.alignment = Alignment(horizontal='center')

        for talhao in terreno.talhoes.all():
            ws.append([
                terreno.cidade,
                terreno.produtor.user.first_name,
                talhao.nome,
                talhao.cidade,
                talhao.data_certificacao.strftime('%d/%m/%Y'),
                talhao.numero_plantas,
                talhao.data_plantio.strftime('%d/%m/%Y'),
                talhao.variedade,
                talhao.area,
                talhao.validade_caf.strftime('%d/%m/%Y')
            ])

        # Ajustar largura das colunas (opcional)
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = column_cells[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_length + 2, 12)

    # ===== GERAR ARQUIVO =====
    nome_arquivo = f"relatorio_produtores_{coordenador.user.first_name}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'
    wb.save(response)
    return response


@api_view(['GET'])
def api_enviar_produtor(request, email):
    chave = request.headers.get('X-API-Key')
    
    # if chave != 'segredo123':
    #    return Response({'erro': 'Acesso não autorizado!'}, status=status.HTTP_401_UNAUTHORIZED)
    
    try:
        produtor = Produtor.objects.select_related('user') \
                                .prefetch_related('terrenos__talhoes') \
                                .get(user__email=email)
        serializer = ProdutorSerializer(produtor)
        return JsonResponse(serializer.data, status=status.HTTP_200_OK)
    except Produtor.DoesNotExist:
        return JsonResponse({'erro': 'Produtor não encontrado!'}, status=status.HTTP_404_NOT_FOUND)

def importar_servicos_view(request):
    produtor = request.GET.get("produtor")
    resultado = importar_servicos_da_api(produtor_email=produtor)
    return JsonResponse(resultado)

def visualizar_servicos(request):
    resultado = importar_servicos_da_api()
    if resultado["success"]:
        servicos = resultado["dados"]
    else:
        servicos = []
        erro = resultado["erro"]
        return render(request, 'coordenador/servicos.html', {'erro': erro})
    
    from datetime import datetime

    # Converte a string da data ISO para objeto datetime
    for s in servicos:
        data_str = s.get("data")
        if isinstance(data_str, str):
            try:
                s["data"] = datetime.strptime(data_str, "%Y-%m-%dT%H:%M:%S.%fZ")
            except ValueError:
                try:
                    # Caso venha sem os milissegundos
                    s["data"] = datetime.strptime(data_str, "%Y-%m-%dT%H:%M:%SZ")
                except ValueError:
                    s["data"] = None


    return render(request, 'coordenador/servicos.html', {'servicos': servicos})

def servicos_por_produtor(request, email):
    resultado = importar_servicos_da_api(produtor_email=email)
    if not resultado["success"]:
        return render(request, 'coordenador/servicos-por-produtor.html', {'erro': resultado["erro"]})

    servicos = resultado["dados"]

    # Converte a string da data ISO para objeto datetime
    for s in servicos:
        data_str = s.get("data")
        if isinstance(data_str, str):
            try:
                s["data"] = datetime.strptime(data_str, "%Y-%m-%dT%H:%M:%S.%fZ")
            except ValueError:
                try:
                    # Caso venha sem os milissegundos
                    s["data"] = datetime.strptime(data_str, "%Y-%m-%dT%H:%M:%SZ")
                except ValueError:
                    s["data"] = None

    return render(request, 'coordenador/servicos-por-produtor.html', {
        'servicos': servicos,
        'email': email
    })
